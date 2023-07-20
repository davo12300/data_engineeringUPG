import openpyxl
import pandas as pd 
import matplotlib.pyplot as ptl


##acceso al documento
doc= openpyxl.load_workbook("guanajuato.xlsx")
##acceso a la hoja
hoja = doc.get_sheet_by_name("Hoja1")
years, tipo, cantidad = [],[],[]

##EXTRACCIÓN DE DATOS 

## años
for row in hoja.iter_rows():
    a = row[6].value
    years.append(a)

## tipo de conteo
for row in hoja.iter_rows():
    b = row[4].value
    tipo.append(b)

## valor del conteo
for row in hoja.iter_rows():
    c = row[7].value
    cantidad.append(c)


## tipo de 

#ELIMINAR PRIMERA FILA

years.remove(years[0])
tipo.remove(tipo[0])
cantidad.remove(cantidad[0])


##definicion de cuentas

years_selected = ['2017','2018','2019','2020','2021']
count_year = []

for x in range(len(years_selected)):
    contador = 0
    for y in range(len(years)):    
        if years[y]==int(years_selected[x]) and tipo[y] ==1006000039:
            contador = contador + cantidad[y]  
    count_year.append(contador)

info = {

    'Años': years_selected,
    'Cantidad de Accidentes':count_year

}
df = pd.DataFrame(info)

print(df)
ptl.bar(df['Años'],df['Cantidad de Accidentes'] , color = '#B0A8B9')
ptl.title('Accidentes de tránsito terrestre en zonas urbanas y suburbanas,  últimos 5 años')
ptl.show()



