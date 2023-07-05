import openpyxl
import pandas as pd 
import matplotlib.pyplot as ptl
from collections import Counter
from tabulate import tabulate

##acceso al documento
doc= openpyxl.load_workbook("salarios.xlsx")
##acceso a la hoja
hoja = doc.get_sheet_by_name("Hoja1")
p = []
s = []
##DATOS 
for row in hoja.iter_rows():
    pesos = row[0].value
    p.append(pesos.upper())
##SALARIOS
for row in hoja.iter_rows():
    money = row[1].value
    s.append(money)

data = [] 
## QUITAR PRIMER DATO DE LA COLUMNA
for x in range(1,len(p)):
    data.append(p[x])

salarios = [] 

for x in range(1,len(s)):
    salarios.append(s[x])


Palabras_correctas = ['WALMART', 'WALDOS','CHEDRAUI','AURRERA','SORIANA','COMERCIAL MEXICANA']


for x in range(len(Palabras_correctas)):
    correct = Counter(Palabras_correctas[x])
    
    

    for y in range(len(data)):
        compara = Counter(data[y])
        info = correct | compara
               
        if correct == info:
            
            data[y]=Palabras_correctas[x]

datos = {
    'INFORMACION': data,
    'SALARIOS':salarios
}
##calculo de salarios jsjs

p_salarios , maxim, index = [], [] ,[]

for n in range(len(Palabras_correctas)):

    suma,c = 0,0
    for z in range(len(data)):
        if Palabras_correctas[n] == data[z]:
            suma = suma + salarios[z]
            c = c + 1


    maxim.append(c)
    
    p_salarios.append(round((suma/c),2))



##ENCONTRAR PRIMEROS 3
while len(index)<=2:
    for d in range(len(maxim)):
        if max(maxim)==maxim[d]:
            
            index.append(d)
            maxim[d]=0
    
index.pop()

##final data frame jsj

infogra = {
    'EMPRESA': [Palabras_correctas[index[0]], Palabras_correctas[index[1]], Palabras_correctas[index[2]]],
    'SALARIO': [p_salarios[index[0]], p_salarios[index[1]],p_salarios[index[2]]],
    
}
dframe = pd.DataFrame(infogra)


print(tabulate(infogra, headers = 'keys', tablefmt = 'pretty'))

ptl.bar(dframe['EMPRESA'],dframe['SALARIO'] , color = ["#007FDD","#54DD00","#DD7200"]
        ,edgecolor ='black' )
ptl.title('SALARIOS POR EMPRESA')
ptl.ylabel('Promedio Salarial')

ptl.show()








