import openpyxl

##acceso al documento
doc= openpyxl.load_workbook("pesos.xlsx")
##acceso a la hoja
hoja = doc.get_sheet_by_name("Hoja1")
p = []

##EXTRACCIÓN DE DATOS 
for row in hoja.iter_rows():
    pesos = row[0].value
    p.append(pesos)
data = [] 
## QUITAR PRIMER DATO DE LA COLUMNA
for x in range(1,len(p)):
    data.append(p[x])


## PROMEDIO
suma = 0
count = 0
for y in range(len(data)):  
    if data[y] is None:
        suma = suma
        count = count + 1  
    else:
        suma = suma + data[y]
promedio= suma/ len(data)

if count == 0:
    print("no hay datos faltantes")

else:
    
    for e in range(len(data)):  
        if data[e] is None:
            data[e] = round(promedio,2)

for row_idx, cell_value in enumerate(data):
    hoja.cell(row=row_idx + 2, column=1, value=cell_value)

    doc.save("pesos.xlsx")



    

